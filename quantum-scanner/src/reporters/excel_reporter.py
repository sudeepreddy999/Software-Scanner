"""
Excel Reporter for OpenSSL Cipher Suite Analysis

Generates simplified Excel report with all cipher suite details in one sheet.
"""

from typing import List, Dict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..openssl_analyzer import CipherSuite, QuantumStrength


class ExcelCipherReporter:
    """Generate Excel reports for OpenSSL cipher suite analysis"""
    
    # Color scheme for quantum strength levels
    STRENGTH_COLORS = {
        'CRITICAL': 'FF0000',      # Red
        'LOW': 'FFA500',           # Orange
        'MEDIUM': 'FFFF00',        # Yellow
        'HIGH': '90EE90',          # Light Green
        'QUANTUM_SAFE': '00FF00'   # Green
    }
    
    def __init__(self):
        self.workbook = None
        
    def generate_report(
        self,
        cipher_suites: List[CipherSuite],
        statistics: Dict,
        openssl_version: str,
        output_file: str
    ):
        """Generate simplified Excel report with all cipher suite details"""
        
        self.workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        # Create single comprehensive sheet
        self._create_all_ciphers_sheet(cipher_suites, openssl_version)
        
        # Save workbook
        self.workbook.save(output_file)
        print(f"Excel report saved to: {output_file}")
    
    def _create_all_ciphers_sheet(self, cipher_suites: List[CipherSuite], openssl_version: str):
        """Create comprehensive sheet with all cipher suite details"""
        
        ws = self.workbook.create_sheet("All Cipher Suites", 0)
        
        # Title and metadata
        ws['A1'] = "OpenSSL Cipher Suite Analysis - Complete Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:L1')
        
        ws['A2'] = f"OpenSSL Version: {openssl_version}"
        ws['A2'].font = Font(size=11, italic=True)
        ws.merge_cells('A2:L2')
        
        ws['A3'] = f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(size=11, italic=True)
        ws.merge_cells('A3:L3')
        
        ws['A4'] = f"Total Cipher Suites: {len(cipher_suites)}"
        ws['A4'].font = Font(size=11, italic=True)
        ws.merge_cells('A4:L4')
        
        # Headers (row 6)
        headers = [
            "Cipher Suite Name",
            "Protocol",
            "Key Exchange",
            "KX Quantum Status",
            "Authentication",
            "Auth Quantum Status",
            "Encryption",
            "Enc Key Size (bits)",
            "Enc Quantum Status",
            "Hash Algorithm",
            "Hash Key Size (bits)",
            "Hash Quantum Status",
            "Overall Quantum Strength",
            "Strength Score",
            "Description",
            "Recommendation"
        ]
        
        header_row = 6
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
        
        # Data rows
        for row_idx, cipher in enumerate(cipher_suites, start=header_row+1):
            ws[f'A{row_idx}'] = cipher.name
            ws[f'B{row_idx}'] = cipher.protocol
            ws[f'C{row_idx}'] = cipher.kx_algorithm
            ws[f'D{row_idx}'] = cipher.kx_quantum_status
            ws[f'E{row_idx}'] = cipher.auth_algorithm
            ws[f'F{row_idx}'] = cipher.auth_quantum_status
            ws[f'G{row_idx}'] = cipher.enc_algorithm
            ws[f'H{row_idx}'] = cipher.enc_key_size if cipher.enc_key_size else "N/A"
            ws[f'I{row_idx}'] = cipher.enc_quantum_status
            ws[f'J{row_idx}'] = cipher.hash_algorithm if cipher.hash_algorithm else "N/A"
            ws[f'K{row_idx}'] = cipher.hash_key_size if cipher.hash_key_size else "N/A"
            ws[f'L{row_idx}'] = cipher.hash_quantum_status
            ws[f'M{row_idx}'] = cipher.quantum_strength.value
            ws[f'N{row_idx}'] = cipher.strength_score
            ws[f'O{row_idx}'] = cipher.description
            ws[f'P{row_idx}'] = cipher.recommendation
            
            # Color code quantum status columns
            quantum_safe_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type="solid")
            quantum_vuln_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type="solid")
            
            # KX status
            kx_cell = ws[f'D{row_idx}']
            kx_cell.fill = quantum_safe_fill if cipher.kx_quantum_status == "Quantum-Safe" else quantum_vuln_fill
            kx_cell.alignment = Alignment(horizontal='center', vertical='center')
            kx_cell.font = Font(bold=True)
            
            # Auth status
            auth_cell = ws[f'F{row_idx}']
            auth_cell.fill = quantum_safe_fill if cipher.auth_quantum_status == "Quantum-Safe" else quantum_vuln_fill
            auth_cell.alignment = Alignment(horizontal='center', vertical='center')
            auth_cell.font = Font(bold=True)
            
            # Enc status
            enc_cell = ws[f'I{row_idx}']
            enc_cell.fill = quantum_safe_fill if cipher.enc_quantum_status == "Quantum-Safe" else quantum_vuln_fill
            enc_cell.alignment = Alignment(horizontal='center', vertical='center')
            enc_cell.font = Font(bold=True)
            
            # Hash status
            hash_cell = ws[f'L{row_idx}']
            hash_cell.fill = quantum_safe_fill if cipher.hash_quantum_status == "Quantum-Safe" else quantum_vuln_fill
            hash_cell.alignment = Alignment(horizontal='center', vertical='center')
            hash_cell.font = Font(bold=True)
            
            # Color code overall quantum strength column
            color = self.STRENGTH_COLORS.get(cipher.quantum_strength.value, 'FFFFFF')
            strength_cell = ws[f'M{row_idx}']
            strength_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            strength_cell.font = Font(bold=True)
            strength_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Format score column
            score_cell = ws[f'N{row_idx}']
            score_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Wrap text for long descriptions
            ws[f'O{row_idx}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws[f'P{row_idx}'].alignment = Alignment(wrap_text=True, vertical='top')
            
            # Add borders to all cells
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 45  # Cipher name
        ws.column_dimensions['B'].width = 12  # Protocol
        ws.column_dimensions['C'].width = 18  # Key Exchange
        ws.column_dimensions['D'].width = 20  # KX Quantum Status
        ws.column_dimensions['E'].width = 18  # Authentication
        ws.column_dimensions['F'].width = 20  # Auth Quantum Status
        ws.column_dimensions['G'].width = 22  # Encryption
        ws.column_dimensions['H'].width = 15  # Enc Key Size
        ws.column_dimensions['I'].width = 20  # Enc Quantum Status
        ws.column_dimensions['J'].width = 15  # Hash Algorithm
        ws.column_dimensions['K'].width = 15  # Hash Key Size
        ws.column_dimensions['L'].width = 20  # Hash Quantum Status
        ws.column_dimensions['M'].width = 20  # Overall Quantum Strength
        ws.column_dimensions['N'].width = 13  # Score
        ws.column_dimensions['O'].width = 50  # Description
        ws.column_dimensions['P'].width = 60  # Recommendation
        
        # Set row heights
        ws.row_dimensions[header_row].height = 40  # Header row
        for row in range(header_row + 1, len(cipher_suites) + header_row + 1):
            ws.row_dimensions[row].height = 35  # Data rows
        
        # Freeze panes (freeze header and metadata)
        ws.freeze_panes = f'A{header_row + 1}'
        
        # Add auto-filter
        ws.auto_filter.ref = f'A{header_row}:P{len(cipher_suites) + header_row}'
